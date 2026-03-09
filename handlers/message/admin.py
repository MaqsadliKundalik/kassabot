from aiogram import Router, F
from aiogram.types import Message, BufferedInputFile
from aiogram.filters import Command
from models.main import InOutFlow, User
from aiogram.fsm.context import FSMContext
from config import KASSA_CHAT_ID, ADMINS, get_tashkent_time
from keyboards import vote_btn, admin_menu
from states import AdminStates
from openpyxl import Workbook
from datetime import datetime
import io
from tortoise.functions import Sum

router = Router()

@router.message(Command("start"))
async def start(message: Message, state: FSMContext):
    await message.answer("Hayrli kun admin!", reply_markup=admin_menu)

@router.message(F.text == "Kassaga pul qo'shish")
async def add_money(message: Message, state: FSMContext):
    await message.answer("Kassaga pul qo'shish uchun summani kiriting:")
    await state.set_state(AdminStates.add_money)

@router.message(AdminStates.add_money)
async def add_money_amount(message: Message, state: FSMContext, user: User):
    # Validate amount input
    amount_text = message.text.strip()
    cleaned_amount = amount_text.replace(' ', '').replace(',', '')
    
    if not cleaned_amount.replace('.', '').isdigit():
        await message.answer("❌ Noto'g'ri format! Iltimos, faqat son kiriting (masalan: 50000, 50 000, 50.500)")
        return
    
    try:
        if '.' in cleaned_amount:
            parts = cleaned_amount.split('.')
            if len(parts) > 2:
                await message.answer("❌ Noto'g'ri format! Iltimos, faqat bitta nuqta ishlating")
                return
            
            whole_part = parts[0] if parts[0] else '0'
            decimal_part = parts[1] if len(parts) > 1 else '0'
            
            if len(decimal_part) > 2:
                amount = float(whole_part + decimal_part)
            else:
                whole = float(whole_part) if whole_part else 0
                tiyin = int(decimal_part.ljust(2, '0')[:2]) / 100
                amount = whole + tiyin
        else:
            amount = float(cleaned_amount)
        
        if amount <= 0:
            await message.answer("❌ Summa musbat son bo'lishi kerak!")
            return
        
        await InOutFlow.create(
            user=user,
            amount=amount,
            type="income",
            description="Kassaga pul qo'shish"
        )
        await message.answer(f"✅ Kassaga {amount} so'm qo'shildi")
        await state.clear()
        
    except ValueError:
        await message.answer("❌ Iltimos, to'g'ri summani kiriting:")

@router.message(F.text == "Kassa hisoboti")
async def kassa_report(message: Message):
    """Kassa hisobotini Excel faylda yuborish"""
    try:
        flows = await InOutFlow.all().prefetch_related('user').order_by('-created_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Kassa Hisoboti"
        
        headers = ["Sana", "Foydalanuvchi", "Turi", "Summa", "Izoh"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row, flow in enumerate(flows, 2):
            ws.cell(row=row, column=1, value=flow.created_at.strftime("%d.%m.%Y %H:%M"))
            ws.cell(row=row, column=2, value=flow.user.name)
            ws.cell(row=row, column=3, value="Kirim" if flow.type == "income" else "Chiqim")
            ws.cell(row=row, column=4, value=float(flow.amount))
            ws.cell(row=row, column=5, value=flow.description)
        
        total_income = sum(f.amount for f in flows if f.type == "income")
        total_outcome = sum(f.amount for f in flows if f.type == "outcome")
        balance = total_income - total_outcome
        
        summary_row = len(flows) + 3
        ws.cell(row=summary_row, column=1, value="Jami kirim:")
        ws.cell(row=summary_row, column=4, value=float(total_income))
        
        ws.cell(row=summary_row + 1, column=1, value="Jami chiqim:")
        ws.cell(row=summary_row + 1, column=4, value=float(total_outcome))
        
        ws.cell(row=summary_row + 2, column=1, value="Qoldiq:")
        ws.cell(row=summary_row + 2, column=4, value=float(balance))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        file_name = f"kassa_hisoboti_{datetime.now().strftime('%d.%m.%Y_%H-%M')}.xlsx"
        document = BufferedInputFile(excel_file.read(), filename=file_name)
        
        await message.answer_document(
            document=document,
            caption=f"📊 Kassa hisoboti\n\nJami kirim: {total_income:,.2f} so'm\nJami chiqim: {total_outcome:,.2f} so'm\nQoldiq: {balance:+,2f} so'm"
        )
        
    except Exception as e:
        await message.answer(f"❌ Hisobotni yaratishda xatolik: {str(e)}")

@router.message(Command("balance"))
async def balance_command(message: Message):
    try:
        sum_income_result = await InOutFlow.filter(type="income").annotate(total=Sum("amount")).first()
        sum_outcome_result = await InOutFlow.filter(type="outcome").annotate(total=Sum("amount")).first()
        
        sum_income = sum_income_result.total if sum_income_result and sum_income_result.total else 0
        sum_outcome = sum_outcome_result.total if sum_outcome_result and sum_outcome_result.total else 0
        balance = sum_income - sum_outcome
        
        await message.answer(f"💰 Balans: {balance:,.2f} so'm\n\n📈 Jami kirim: {sum_income:,.2f} so'm\n📉 Jami chiqim: {sum_outcome:,.2f} so'm")
    except Exception as e:
        await message.answer(f"❌ Balansni hisoblashda xatolik: {str(e)}")